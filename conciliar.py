"""
EMETE Asesoría — Conciliador Bancario Cegid

Ejecutar la conciliación real: ajustar las rutas en CONFIGURACIÓN y `python conciliar.py`.
Ejecutar sin los ficheros presentes: `python conciliar.py` -> imprime "self-check OK"
(sirve para verificar la instalación, igual que trimestre.py del organizador).

Requiere: pip install -r requirements.txt   (openpyxl, pandas, python-dateutil)
La API de Anthropic es opcional (solo si se define API_KEY): pip install anthropic
"""
import re, json, os, time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

# ── CONFIGURACIÓN (ajustar por ejecución) ────────────────────────────────────
EXTRACTO_PATH = "extracto.xlsx"     # ruta al extracto bancario
LM_PATH       = "mayor.xlsx"        # ruta al libro mayor de Cegid
PGC_PATH      = "plan_cuentas.xlsx" # ruta al plan de cuentas de Cegid
OUTPUT_PATH   = "conciliado.xlsx"   # fichero de salida
API_KEY       = ""                  # API key Anthropic (opcional)
FECHA_DESDE   = None                # "2026-01-01" o None para todo
FECHA_HASTA   = None                # "2026-03-31" o None para todo

# ── DETECCIÓN AUTOMÁTICA DE FORMATO ──────────────────────────────────────────
_KW = {
    'F_CONTABLE':    ['contable','fecha op','f. contable','fecha valor','date','fecha'],
    'CONCEPTO':      ['concepto','descripcion','descripción','motivo','concept','detalle'],
    'BENEFICIARIO':  ['beneficiario','ordenante','nombre','benef'],
    'OBSERVACIONES': ['observaciones','observacion','notas','info','adicional'],
    'IMPORTE':       ['importe','cantidad','amount','cargo','abono','movimiento'],
    'SALDO':         ['saldo','balance','disponible'],
    'F_VALOR':       ['valor','f. valor','value date'],
    'CODIGO':        ['código','codigo','code','referencia'],
}

def _detectar_formato(ws):
    for fi, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        celdas = [(i, str(v).strip().lower()) for i, v in enumerate(row) if v and str(v).strip()]
        if len(celdas) < 2: continue
        mapa = {}
        for campo, kws in _KW.items():
            for i, texto in celdas:
                if any(kw in texto for kw in kws) and campo not in mapa:
                    mapa[campo] = i; break
        if 'F_CONTABLE' in mapa and 'IMPORTE' in mapa:
            col_off = min(mapa.values())
            return fi + 1, col_off, mapa
    # Fallback BBVA
    return 17, 2, {'F_CONTABLE':2,'F_VALOR':3,'CODIGO':4,'CONCEPTO':5,
                   'BENEFICIARIO':6,'OBSERVACIONES':7,'IMPORTE':8,'SALDO':9}

def _smart_float(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    s = str(val).replace('EUR','').replace('€','').strip()
    if '.' in s and ',' in s:
        if s.rindex('.') < s.rindex(','):
            s = s.replace('.','').replace(',','.')
        else:
            s = s.replace(',','')
    elif ',' in s:
        s = s.replace(',','.')
    s = re.sub(r'[^\d\.\-]','',s)
    try: return float(s)
    except: return 0.0

_NUM_PDF = re.compile(
    r'^(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+(\d{4,5})'
    r'(?:\s+(.+?))??\s+(-?[\d\.]*,\d{1,2})\s+EUR\s+(-?[\d\.]*,\d{1,2})\s+EUR\s*$')
_SKIP_PDF = ('Movimientos','Titular ','Cuenta ES','Divisa ','Banco BANCO','F. CONTABLE','F.CONTABLE')

def _norm_pdf(t):
    import unicodedata
    if not t: return ''
    t = unicodedata.normalize('NFD', str(t)).encode('ascii','ignore').decode('ascii')
    return re.sub(r'\s+', ' ', t).strip()

def _parsear_movimientos_pdf(lineas):
    """lineas: list[(page, top, text)] ya filtradas. Devuelve list[dict] estándar."""
    # clustering por salto vertical (>18 = nuevo movimiento); corta también al cambiar de página
    movs, cur, pp, pt = [], [], None, None
    for pi, t, txt in lineas:
        if cur and (pi != pp or pt is None or t - pt > 18):
            movs.append(cur); cur = []
        cur.append((t, txt)); pp, pt = pi, t
    if cur: movs.append(cur)
    rows = []
    for mv in movs:
        ni = m = None
        for i, (_, txt) in enumerate(mv):
            mm = _NUM_PDF.match(txt)
            if mm: ni, m = i, mm; break
        if ni is None: continue
        fc, fv, cod, cmid, imp, saldo = m.groups()
        head = _norm_pdf(' '.join(txt for _, txt in mv[:ni]) + (' ' + cmid if cmid else ''))
        obs = _norm_pdf(' '.join(txt for _, txt in mv[ni+1:]))
        conc, benef = (head.split('|', 1) + [''])[:2] if '|' in head else (head, '')
        rows.append({'F_CONTABLE': fc, 'F_VALOR': fv, 'CODIGO': cod,
                     'CONCEPTO': _norm_pdf(conc), 'BENEFICIARIO': _norm_pdf(benef),
                     'OBSERVACIONES': obs, 'IMPORTE': _smart_float(imp),
                     'SALDO': _smart_float(saldo)})
    return rows

def cargar_extracto_pdf(path):
    import pdfplumber
    from collections import defaultdict
    lineas = []
    with pdfplumber.open(path) as pdf:
        for pi, pg in enumerate(pdf.pages):
            grp = defaultdict(list)
            for w in pg.extract_words(use_text_flow=False, keep_blank_chars=False):
                grp[round(w['top'])].append(w)
            for t in sorted(grp):
                txt = ' '.join(w['text'] for w in sorted(grp[t], key=lambda x: x['x0']))
                if re.match(r'^\d+/\d+$', txt): continue
                if any(txt.startswith(s) for s in _SKIP_PDF): continue
                lineas.append((pi, t, txt))
    rows = _parsear_movimientos_pdf(lineas)
    df = pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=['F_CONTABLE','F_VALOR','CODIGO','CONCEPTO','BENEFICIARIO',
                 'OBSERVACIONES','IMPORTE','SALDO'])
    roturas = validar_saldo(df)
    if roturas:
        print(f"AVISO: {len(roturas)} posible(s) rotura(s) de saldo en el PDF (filas {roturas[:10]}...)")
    return df

def cargar_extracto(path, fecha_desde=None, fecha_hasta=None):
    if path.lower().endswith('.pdf'):
        return cargar_extracto_pdf(path)   # el filtro de fechas se aplica aguas arriba si hace falta
    import unicodedata, shutil
    # Si el fichero es CSV pero contiene XLSX (CaixaBank), renombrar
    tmp = path
    if path.lower().endswith('.csv'):
        with open(path,'rb') as f:
            if f.read(2) == b'PK':
                tmp = path + '.xlsx'
                shutil.copy(path, tmp)
    wb = load_workbook(tmp, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    fila_datos, col_off, mapa = _detectar_formato(ws)
    mapa_rel = {k: v - col_off for k, v in mapa.items()}
    rows = []
    for row in ws.iter_rows(min_row=fila_datos, values_only=True):
        v = row[col_off:] if col_off < len(row) else ()
        if not v: continue
        def _safe(idx, default=''):
            try: return v[idx] if v[idx] is not None else default
            except IndexError: return default
        fecha_raw = _safe(mapa_rel.get('F_CONTABLE',0))
        if fecha_raw in (None,'','None','nan'): continue
        # Filtro de fechas
        if fecha_desde or fecha_hasta:
            try:
                from datetime import date as _date
                if hasattr(fecha_raw, 'date'):
                    fd = fecha_raw.date()
                else:
                    from dateutil.parser import parse
                    fd = parse(str(fecha_raw)).date()
                if fecha_desde and fd < fecha_desde: continue
                if fecha_hasta and fd > fecha_hasta: continue
                fecha_str = fd.strftime('%d/%m/%Y')
            except: fecha_str = str(fecha_raw)
        else:
            fecha_str = str(fecha_raw).strip()
        imp = _smart_float(_safe(mapa_rel.get('IMPORTE',6)))
        if imp == 0.0 and 'IMPORTE' not in mapa_rel: continue
        def norm(t):
            if not t or t in ('None','nan',''): return ''
            t = unicodedata.normalize('NFD', str(t)).encode('ascii','ignore').decode('ascii')
            return t.strip()
        rows.append({
            'F_CONTABLE':    fecha_str,
            'F_VALOR':       norm(_safe(mapa_rel.get('F_VALOR',1))),
            'CODIGO':        norm(_safe(mapa_rel.get('CODIGO',2))),
            'CONCEPTO':      norm(_safe(mapa_rel.get('CONCEPTO',3))),
            'BENEFICIARIO':  norm(_safe(mapa_rel.get('BENEFICIARIO',4))),
            'OBSERVACIONES': norm(_safe(mapa_rel.get('OBSERVACIONES',5))),
            'IMPORTE':       imp,
            'SALDO':         _smart_float(_safe(mapa_rel.get('SALDO',7))),
        })
    wb.close()
    if tmp != path and os.path.exists(tmp): os.remove(tmp)
    return pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=['F_CONTABLE','F_VALOR','CODIGO','CONCEPTO',
                 'BENEFICIARIO','OBSERVACIONES','IMPORTE','SALDO'])

def cargar_lm(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    terceros = {}
    importes = {}
    cod = None
    for row in ws.iter_rows(min_row=5, values_only=True):
        a = str(row[0]).strip() if row[0] else ''
        b = row[1]
        if a and b is None and a not in ['','None']:
            if not any(x in a for x in ['Saldo','Total','periodo','ejercicio']):
                p = a.split(' ',1); c = p[0].strip()
                nom = p[1].strip() if len(p)>1 else a
                if c[:2] in ('40','41','43'):
                    cod = c; terceros[cod] = nom; importes.setdefault(cod, set())
                else:
                    cod = None
        elif cod:
            # fila de apunte del tercero actual: guardar importes numéricos > 0
            for cell in row[1:]:
                if isinstance(cell, (int, float)) and abs(cell) > 0:
                    importes[cod].add(round(abs(float(cell)), 2))
    wb.close()
    init_importes_lm(importes)
    return terceros

def cargar_pgc(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Datos"]
    pgc = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        if row[4] and row[9]:
            pgc[str(row[4]).strip()] = str(row[9]).strip()
        # Task 1: indexar también la subcuenta de 8 dígitos (columna 8)
        if len(row) > 8 and row[8] and row[9]:
            pgc[str(row[8]).strip()] = str(row[9]).strip()
    wb.close()
    return pgc

def cargar_criterios(carpeta):
    """Lee el primer criterios*.json de la carpeta. {} si no hay."""
    import glob
    try:
        matches = sorted(glob.glob(os.path.join(carpeta, 'criterios*.json')))
    except Exception:
        return {}
    if not matches:
        return {}
    try:
        with open(matches[0], encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"criterios.json ilegible: {e}")
        return {}

# ── MATCHING LIBRO MAYOR ──────────────────────────────────────────────────────
STOPWORDS = {'de','del','la','el','los','las','y','e','a','en','sl','slu','sa',
             'sau','ltd','srl','slp','sucursal','espana','spain','iberia',
             'soc','cia','grupo','proyectos','servicios','soluciones','espa'}

def toks(t):
    if not t: return set()
    import unicodedata
    t = unicodedata.normalize('NFD', str(t)).encode('ascii','ignore').decode('ascii')
    # Task 2: separar minúscula→MAYÚSCULA pegadas (camelCase) antes de subir todo a mayúsculas
    t = re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', t)
    t = re.sub(r'[^A-Z0-9\s]',' ', t.upper())
    return {x for x in t.split() if len(x)>2 and x.lower() not in STOPWORDS}

_cache_lm = {}
def init_cache_lm(terceros):
    global _cache_lm
    _cache_lm = {cod: toks(nom) for cod, nom in terceros.items()}

_importes_lm = {}
def init_importes_lm(mapa):
    global _importes_lm
    _importes_lm = {c: {round(abs(x), 2) for x in s} for c, s in mapa.items()}

def match_importe(imp):
    """Devuelve el cod de tercero cuyo LM tiene un apunte del mismo importe, si es único."""
    objetivo = round(abs(float(imp)), 2)
    candidatos = [c for c, s in _importes_lm.items()
                  if any(abs(v - objetivo) <= 0.02 for v in s)]
    return candidatos[0] if len(candidatos) == 1 else None

def match_lm(beneficiario, observaciones, concepto, terceros):
    mejor, bcod, bnom, bcampo = 0.0, None, None, ''
    for texto, peso, campo in [(beneficiario,1.1,'benef'),(observaciones,1.0,'obs'),(concepto,0.85,'conc')]:
        te = toks(texto)
        if not te: continue
        for cod, tt in _cache_lm.items():
            if not tt: continue
            inter = te & tt
            if not inter: continue
            score = min((len(inter)/len(tt))*peso, 1.0)
            if score > mejor:
                mejor, bcod, bnom, bcampo = score, cod, terceros[cod], campo
    return (bcod, bnom, mejor, bcampo) if mejor >= 0.6 else None

# ── REGLAS DE CLASIFICACIÓN ───────────────────────────────────────────────────
R_OBS = [
    # Nóminas: permitir texto entre NOMINA y el mes (Task 5)
    (r'NOMIN[A-Z]*\b.{0,25}(ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE)',
     '46500000','REMUNERACIONES PENDIENTES DE PAGO','ALTA'),
    (r'(PAGO|ADELANTO|ANTICIPO)\s+.{0,10}NOMIN|NOMIN.{0,10}(PAGO|ADELANTO|ANTICIPO)',
     '46500000','REMUNERACIONES PENDIENTES DE PAGO','ALTA'),
    # Cotización SS: aceptar 'SEGURID' truncado (Task 5)
    (r'TGSS|TESORERIA\s+GENERAL.{0,20}SEGURID|COTIZACION\s+SS|CUOTA\s+SS|SEG\.?\s*SOCIAL',
     '47600000','ORGANISMOS SEGURIDAD SOCIAL, ACREEDORES','ALTA'),
    # TPV / liquidación remesa de comercios (cuenta destino por criterios) (Task 5)
    (r'LIQUIDACI[OO]N\s+REMESA\s+DE\s+COMERCIOS',
     '43000000','VENTAS TPV (segun cliente)','MEDIA'),
    # Préstamos / devoluciones socios
    (r'PRESTAMO\s+DE\s+SOC|PR[EE]STAMO\s+SOC',
     '55100000','C/C SOCIOS Y ADMINISTRADORES','ALTA'),
    (r'DEVOLUCI[OO]N?\s+PR[EE]?STAMO',
     '55100000','C/C SOCIOS Y ADMINISTRADORES','ALTA'),
    # Traspaso entre cuentas propias
    (r'TRANSFERENCIA\s+ENTRE\s+CUENTAS|TRASPASO\s+ENTRE\s+CUENTAS|\bTRASPASO\b',
     '57200000','BANCOS E INSTITUCIONES DE CREDITO','ALTA'),
    # Pago de impuesto con NRC
    (r'\bNRC[\.\s:\-]',
     '47500000','H.P. ACREEDORA POR CONCEPTOS FISCALES','ALTA'),
    # Compra de vehículo
    (r'COMPRA\s+COCHE|COMPRA\s+VEH[II]CULO|COMPRA\s+FURG|COMPRA\s+RENAULT|COMPRA\s+DACIA',
     '21800000','ELEMENTOS DE TRANSPORTE','REVISAR'),
    # PRL
    (r'PRL\b|PREVENCION\s+RIESGO|RECONOCIMIENTO\s+MEDICO|SERVICIO\s+PREVENCI|EXAMEN\s+MEDICO',
     '41000001','PREVENTO (o proveedor PRL del cliente)','ALTA'),
    # Devolución préstamo corto (CaixaBank: "Dev.Prest.")
    (r'DEV\.?\s*PREST|DEVOLUCION\s+PREST',
     '55100000','C/C SOCIOS Y ADMINISTRADORES','ALTA'),
]

R_EXCEPCIONES = [
    # Comisiones/intereses bancarios propios
    (r'LIQUIDACI[OO]N\s+DE\s+INTERES|COMISI[OO]N|COMISIONES?\s+(POR\s+SERV|SERVICIOS?\s+TELEMATIC)'
     r'|MANTENIMIENTO\s+CUENTA|P\.SERV\.\s+TRF|V\.NEGOCIOS\s+CRED|PRECIO\s+ED\.\s+EXTRACTO'
     r'|CUOTA\s+T\.\s+V|ADMINISTRACI[OO]N\s+DEP',
     '62600000','SERVICIOS BANCARIOS Y SIMILARES','MEDIA',
     'Comision/mantenimiento/servicio bancario propio'),
    # Impuestos
    (r'CARGO\s+POR\s+PAGO\s+DE\s+IMPUESTO|PAGO\s+CON\s+TARJETA\s+DE\s+TASA|\bTRIBUTO'
     r'|\bI\.?V\.?A\.?\b|\bI\.?R\.?P\.?F\.?\b|MOD(?:ELO)?[\s\.]+(?:303|111|115|130|200|202)'
     r'|AEAT|AGENCIA\s+TRIBUTARIA|IMPUESTOS?\s+AEAT',
     '47500000','H.P. ACREEDORA POR CONCEPTOS FISCALES','MEDIA',
     'Pago impuesto/tributo AEAT'),
    # ONG / donativos
    (r'FUNDACI[OO]N|ADEUDO\s+DE\s+FUNDACI|ONG\b|O\.N\.G|MEDICOS\s+SIN\s+FRONTERAS|CRUZ\s+ROJA',
     '67800000','GASTOS EXTRAORDINARIOS - DONATIVO (NO DEDUCIBLE IS)','MEDIA',
     'Donativo ONG — no deducible IS'),
    # Cashback / bonificaciones banco
    (r'CASHBACK|RETENCI[OO]N\s+PROMOCI[OO]N|BONIF\.\s+DEVOLUCION',
     '76900000','OTROS INGRESOS FINANCIEROS','MEDIA',
     'Cashback/bonificacion tarjeta'),
    # Retirada cajero → banco a caja
    (r'RET[\.\s]+EFECTIVO|RETIRADA\s+EFECTIVO|CAJERO\s+AUT|REINT\.CAJERO',
     '57000000','CAJA, EUROS','ALTA',
     'Retirada efectivo cajero: 572->570'),
    # Cheque ingresado
    (r'ABONO\s+POR\s+CHEQUE\s+INGRESADO',
     '57100000','CHEQUES','MEDIA',
     'Cheque ingresado en cuenta'),
    # Transferencia recibida sin identificar → revisar
    (r'TRANSF\.\s+A\s+SU\s+FAVOR|TRANSFER\s+INMEDIATA|TRANSFERENCIA\s+RECIBIDA',
     '43000000','CLIENTES (identificar ordenante)','REVISAR',
     'Transferencia recibida — REVISAR: identificar cliente remitente'),
    # Anulaciones
    (r'ANULACI[OO]N\s+ADEUDO|ANULACI[OO]N\s+OPERACI',
     'PENDIENTE','ANULACION — usar contrapartida del apunte original','REVISAR',
     'Anulacion de adeudo'),
    # Seguros domiciliados
    (r'SEGURCAIXA|SEGURCA|PREVISION\s+BALEAR',
     '62500000','PRIMAS DE SEGUROS','MEDIA',
     'Seguro domiciliado — sin factura separada'),
    # Restaurantes / representación
    (r'RESTAURANTE|CAFETERIA|CAFETERIA|HOSTELERIA',
     '62900000','OTROS SERVICIOS (REPRESENTACION)','MEDIA',
     'Gasto representacion'),
]

SOCIOS = []  # adaptar por cliente: apellidos/nombres de socios y administradores

def extraer_proveedor(obs):
    obs = re.sub(r'\*+\d{4}\s*','', str(obs))
    obs = re.sub(r'\d{16}\s*','', obs).strip()
    return obs[:50] if len(obs)>3 else ''

def clasificar_fila(row, terceros, criterios=None):
    criterios = criterios or {}
    conc   = str(row['CONCEPTO']).upper()
    obs    = str(row['OBSERVACIONES']).upper()
    obs_r  = str(row['OBSERVACIONES'])
    benef  = str(row['BENEFICIARIO']).upper()
    benef_r= str(row['BENEFICIARIO'])
    imp    = float(row['IMPORTE'])
    texto_obs = f"{obs} {conc}"
    # 0. Alias marca→cuenta (criterios por cliente): gana a todo lo demás
    texto_alias = f"{conc} {obs} {benef}"
    for marca, cuenta in (criterios.get('alias_terceros') or {}).items():
        if marca.upper() in texto_alias:
            nom = terceros.get(cuenta, '')
            return cuenta, nom, 'ALTA', f"Alias criterios: «{marca}» → {cuenta}", 'ALIAS'
    # 1. Empresa propia
    # NOTA: actualizar con el nombre/CIF de la empresa del cliente
    # 2. Observaciones inequívocas (ANTES que socios)
    for pat, cta, desc, conf in R_OBS:
        if re.search(pat, texto_obs, re.IGNORECASE):
            # overrides por criterios de cliente (Task 5)
            if cta == '43000000' and criterios.get('cuenta_ventas_tpv'):
                cta = criterios['cuenta_ventas_tpv']
            if cta == '41000001' and criterios.get('cuenta_prl'):
                cta = criterios['cuenta_prl']
            fuente = obs_r.strip() or conc[:50]
            return cta, desc, conf, f"Detectado en: «{fuente}»", 'REGLA_OBS'
    # 2b. Ingreso en efectivo → 551
    if re.search(r'INGRESO\s+EN\s+EFECTIVO', conc, re.IGNORECASE) and imp > 0:
        return '55100000','C/C SOCIOS Y ADMINISTRADORES','REVISAR',\
               'Ingreso efectivo — REVISAR: identificar origen','REGLA_OBS'
    # 3. Socios conocidos del cliente (lista adaptar por cliente)
    if any(s in benef for s in SOCIOS):
        return '55100000','C/C SOCIOS Y ADMINISTRADORES','REVISAR',\
               'Transferencia a socio sin referencia — REVISAR','SOCIO'
    # 3b. ONG antes del LM
    if re.search(r'FUNDACI[OO]N|ADEUDO\s+DE\s+FUNDACI|ONG\b|MEDICOS\s+SIN\s+FRONTERAS',
                 texto_obs, re.IGNORECASE):
        return '67800000','GASTOS EXTRAORDINARIOS - DONATIVO (NO DEDUCIBLE IS)','MEDIA',\
               f'Donativo ONG — no deducible IS','REGLA_OBS'
    # 4. Libro Mayor
    m = match_lm(row['BENEFICIARIO'], row['OBSERVACIONES'], row['CONCEPTO'], terceros)
    if m:
        cod_lm, nom_lm, score, campo = m
        conf = 'ALTA' if score >= 0.7 else 'MEDIA'
        return cod_lm, nom_lm, conf, \
               f"Match LM ({campo}): «{nom_lm}» score {score:.2f}", 'LIBRO_MAYOR'
    # 4b. Match por importe contra el LM (solo si criterios no lo desactiva)
    if criterios.get('match_por_importe', True):
        cod_imp = match_importe(imp)
        if cod_imp:
            return cod_imp, terceros.get(cod_imp, ''), 'ALTA', \
                   f"Match por importe: {abs(imp):.2f} → {cod_imp}", 'MATCH_IMPORTE'
    # 5. Excepciones al 41000000
    texto = f"{conc} {obs} {benef}"
    for pat, cta, desc, conf, just in R_EXCEPCIONES:
        if re.search(pat, texto, re.IGNORECASE):
            # Abonos positivos en comisiones → ingreso, no gasto
            if imp > 0 and cta == '62600000':
                return '76900000','OTROS INGRESOS FINANCIEROS','MEDIA',\
                       'Abono comision bancaria → ingreso financiero','REGLA_EXCEPCION'
            return cta, desc, conf, just, 'REGLA_EXCEPCION'
    # 6. Sin match → acreedor genérico
    prov = extraer_proveedor(row['OBSERVACIONES']) or benef_r.strip() or conc[:40]
    return '41000000','ACREEDORES POR PRESTACIONES DE SERVICIOS','BAJA',\
           f'Sin proveedor en LM — crear tercero | {prov}', 'ACREEDOR_GENERICO'

def conciliar(df, terceros, pgc, criterios=None, api_key='', lote=30):
    init_cache_lm(terceros)
    resultados = []
    for idx, row in df.iterrows():
        cta, desc, conf, just, met = clasificar_fila(row, terceros, criterios)
        resultados.append({
            'idx': idx, 'CUENTA': cta, 'DESCRIPCION': desc,
            'CONFIANZA': conf, 'JUSTIFICACION': just, 'METODO': met
        })
    # Claude API para ACREEDOR_GENERICO si hay API key
    if api_key:
        try:
            import anthropic
            pendientes = [r for r in resultados if r['METODO']=='ACREEDOR_GENERICO']
            if pendientes:
                client = anthropic.Anthropic(api_key=api_key)
                ctx = '\n'.join(f"{k}: {v}" for k,v in list(pgc.items())[:200])
                for i in range(0, len(pendientes), lote):
                    bl = pendientes[i:i+lote]
                    payload = json.dumps([{
                        'idx':r['idx'],
                        'concepto':df.iloc[r['idx']]['CONCEPTO'],
                        'beneficiario':df.iloc[r['idx']]['BENEFICIARIO'],
                        'observaciones':df.iloc[r['idx']]['OBSERVACIONES'],
                        'importe':df.iloc[r['idx']]['IMPORTE'],
                    } for r in bl], ensure_ascii=False)
                    resp = client.messages.create(
                        model="claude-sonnet-4-6", max_tokens=2000,
                        system="Contable senior PGC 2007. Pagos a proveedor sin LM → 41000000. "
                               "Solo comisiones banco propias → 62600000. "
                               "Responde SOLO JSON array: {\"idx\":N,\"cuenta\":\"X\",\"descripcion\":\"Y\","
                               "\"confianza\":\"ALTA|MEDIA|BAJA|REVISAR\",\"justificacion\":\"Z\"}",
                        messages=[{"role":"user","content":f"PGC:\n{ctx}\n\nMovimientos:\n{payload}"}]
                    )
                    txt = re.sub(r'^```json\s*|\s*```$','', resp.content[0].text.strip())
                    for cr in json.loads(txt):
                        for r in resultados:
                            if r['idx']==cr['idx'] and r['METODO']=='ACREEDOR_GENERICO':
                                r.update({'CUENTA':cr.get('cuenta','41000000'),
                                          'DESCRIPCION':cr.get('descripcion',''),
                                          'CONFIANZA':cr.get('confianza','BAJA'),
                                          'JUSTIFICACION':cr.get('justificacion',''),
                                          'METODO':'CLAUDE'})
                    if i+lote < len(pendientes): time.sleep(0.3)
        except Exception as e:
            print(f"Claude API: {e}")
    marcar_retrocesiones(df, resultados)
    resultados.sort(key=lambda x: x['idx'])
    return resultados

def marcar_retrocesiones(df, resultados):
    """Empareja movimientos de importe opuesto y fecha próxima → 555 (netean a 0)."""
    from datetime import datetime as _dt
    def _fecha(s):
        for fmt in ('%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
            try: return _dt.strptime(str(s).strip()[:19], fmt)
            except Exception: pass
        return None
    res_map = {r['idx']: r for r in resultados}
    usados = set()
    filas = list(df.iterrows())
    for a in range(len(filas)):
        ia, ra = filas[a]
        if ia in usados: continue
        fa = _fecha(ra['F_CONTABLE']); impa = round(float(ra['IMPORTE']), 2)
        if impa == 0: continue
        for b in range(a+1, len(filas)):
            ib, rb = filas[b]
            if ib in usados: continue
            if round(float(rb['IMPORTE']), 2) != -impa: continue
            fb = _fecha(rb['F_CONTABLE'])
            if fa and fb and abs((fb - fa).days) > 5: continue
            for idx in (ia, ib):
                r = res_map.get(idx)
                if r:
                    r.update({'CUENTA':'55500000',
                              'DESCRIPCION':'PARTIDAS PENDIENTES DE APLICACION',
                              'CONFIANZA':'MEDIA',
                              'JUSTIFICACION':'Par +/- retrocedido (netea a 0)',
                              'METODO':'RETROCESION'})
            usados.add(ia); usados.add(ib)
            break

def validar_saldo(df):
    """Índices i donde saldo[i] != saldo[i+1] + importe[i] (orden descendente, tol 0,02)."""
    roturas = []
    for i in range(len(df) - 1):
        esperado = float(df['SALDO'].iloc[i+1]) + float(df['IMPORTE'].iloc[i])
        if abs(esperado - float(df['SALDO'].iloc[i])) > 0.02:
            roturas.append(i)
    return roturas

def siguiente_codigo_libre(terceros, bloque):
    """bloque: '400'|'410'|'430'. Siguiente subcuenta de 8 díg libre en ese bloque."""
    usados = [int(c) for c in terceros
              if c.isdigit() and len(c) == 8 and c.startswith(bloque)]
    base = int(bloque + '00000')  # p.ej. 40000000
    return str((max(usados) + 1) if usados else base + 1)

def detectar_acciones(df, resultados, terceros, pgc):
    """Terceros a crear (genéricos agrupados), subcuentas inexistentes y avisos."""
    pgc = pgc or {}
    acc = {'terceros_crear': [], 'subcuentas_crear': [], 'avisos': []}
    # 1) terceros a crear: agrupar los ACREEDOR_GENERICO por nombre normalizado
    grupos = {}
    for r in resultados:
        if r['METODO'] != 'ACREEDOR_GENERICO':
            continue
        row = df.iloc[r['idx']]
        nombre = (str(row['BENEFICIARIO']).strip() or str(row['OBSERVACIONES']).strip()
                  or str(row['CONCEPTO']).strip())[:50]
        clave = ' '.join(sorted(toks(nombre))) or nombre.upper()
        grupos.setdefault(clave, {'nombre': nombre, 'n': 0, 'importe': 0.0})
        grupos[clave]['n'] += 1
        grupos[clave]['importe'] += float(row['IMPORTE'])
    reservados = dict(terceros)
    for g in grupos.values():
        bloque = '410' if g['importe'] < 0 else '430'  # pagos→acreedor, cobros→cliente
        cod = siguiente_codigo_libre(reservados, bloque)
        reservados[cod] = g['nombre']
        acc['terceros_crear'].append({
            'codigo': cod, 'nombre': g['nombre'],
            'tipo': 'acreedor' if bloque == '410' else 'cliente',
            'nif': '(pendiente)', 'n_movimientos': g['n']})
    # 2) subcuentas de contrapartida que no existen en el plan
    vistos = set()
    for r in resultados:
        cta = r['CUENTA']
        if cta in ('PENDIENTE',) or not str(cta).isdigit():
            continue
        if cta in vistos or cta in pgc:
            continue
        vistos.add(cta)
        acc['subcuentas_crear'].append({'codigo': cta, 'nombre': r['DESCRIPCION']})
    # 3) avisos: movimientos sin nombre/NIF y roturas de saldo
    for r in resultados:
        row = df.iloc[r['idx']]
        if r['METODO'] == 'ACREEDOR_GENERICO' and not str(row['BENEFICIARIO']).strip() \
           and not str(row['OBSERVACIONES']).strip():
            acc['avisos'].append({'idx': r['idx'], 'motivo': 'Sin beneficiario ni observaciones',
                                  'importe': float(row['IMPORTE'])})
    for i in validar_saldo(df):
        acc['avisos'].append({'idx': i, 'motivo': 'Posible rotura de saldo', 'importe': None})
    return acc

def _hoja_acciones(wb, acciones):
    if 'ACCIONES_CEGID' in wb.sheetnames: del wb['ACCIONES_CEGID']
    aw = wb.create_sheet('ACCIONES_CEGID')
    aw.append(['ANTES DE IMPORTAR EN CEGID — hacer en este orden'])
    aw.append([])
    aw.append(['1) TERCEROS A CREAR'])
    aw.append(['Codigo', 'Nombre', 'Tipo', 'NIF', 'Nº movs'])
    for t in acciones.get('terceros_crear', []):
        aw.append([t['codigo'], t['nombre'], t['tipo'], t['nif'], t['n_movimientos']])
    aw.append([])
    aw.append(['2) SUBCUENTAS A CREAR (no existen en el plan)'])
    aw.append(['Codigo', 'Nombre'])
    for s in acciones.get('subcuentas_crear', []):
        aw.append([s['codigo'], s['nombre']])
    aw.append([])
    aw.append(['3) AVISOS (revisar antes de importar)'])
    aw.append(['Fila', 'Motivo', 'Importe'])
    for a in acciones.get('avisos', []):
        aw.append([a['idx'], a['motivo'], a['importe']])
    for col, w in zip('ABCDE', (16, 40, 12, 16, 10)):
        aw.column_dimensions[col].width = w

def _hoja_cegid(wb, df, resultados):
    if 'CEGID' in wb.sheetnames: del wb['CEGID']
    cg = wb.create_sheet('CEGID')
    cg.append(['Fecha', 'Concepto', 'Importe', 'Contrapartida'])
    res_map = {r['idx']: r for r in resultados}
    for idx in range(len(df)):
        r = res_map.get(idx)
        if not r: continue
        row = df.iloc[idx]
        cg.append([row['F_CONTABLE'], row['CONCEPTO'], row['IMPORTE'], r['CUENTA']])
    for col, w in zip('ABCD', (14, 50, 14, 16)):
        cg.column_dimensions[col].width = w

def generar_excel(df, resultados, extracto_path, output_path, acciones=None):
    # Intentar cargar el original; si falla, crear nuevo
    try:
        wb = load_workbook(extracto_path)
        ws = wb[wb.sheetnames[0]]
        # Detectar la fila de cabecera para añadir columnas
        _, _, mapa = _detectar_formato(ws)
        fila_cab = None
        fila_datos = None
        col_off = min(mapa.values())
        # Buscar fila con cabeceras
        for fi, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
            celdas = [(i, str(v).strip().lower()) for i,v in enumerate(row) if v]
            if any('concepto' in c[1] or 'fecha' in c[1] for c in celdas):
                fila_cab = fi
                fila_datos = fi + 1
                break
        if not fila_cab:
            fila_cab = 1; fila_datos = 2
        C1 = ws.max_column + 1
    except Exception:
        # Crear workbook nuevo si el original no es compatible
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extracto"
        fila_cab = 1; fila_datos = 2; C1 = 9
        # Cabecera original
        ws.append(['F_CONTABLE','F_VALOR','CODIGO','CONCEPTO','BENEFICIARIO',
                   'OBSERVACIONES','IMPORTE','SALDO'])
        for idx, row in df.iterrows():
            ws.append([row['F_CONTABLE'],row['F_VALOR'],row['CODIGO'],row['CONCEPTO'],
                       row['BENEFICIARIO'],row['OBSERVACIONES'],row['IMPORTE'],row['SALDO']])
    CABS = ['CUENTA_CONTRAPARTIDA','DESCRIPCION_CUENTA','CONFIANZA','JUSTIFICACION','METODO']
    fc = PatternFill('solid', fgColor='1F3864')
    ff = Font(bold=True, color='FFFFFF', size=9)
    for i, cab in enumerate(CABS):
        c = ws.cell(row=fila_cab, column=C1+i, value=cab)
        c.font, c.fill = ff, fc
        c.alignment = Alignment(horizontal='center')
    FILLS = {'ALTA':PatternFill('solid',fgColor='C6EFCE'),
             'MEDIA':PatternFill('solid',fgColor='FFEB9C'),
             'BAJA':PatternFill('solid',fgColor='FFCC99'),
             'REVISAR':PatternFill('solid',fgColor='FFC7CE')}
    res_map = {r['idx']:r for r in resultados}
    for idx in range(len(df)):
        fila = fila_datos + idx
        r = res_map.get(idx)
        if not r: continue
        for j, col in enumerate(['CUENTA','DESCRIPCION','CONFIANZA','JUSTIFICACION','METODO']):
            ws.cell(fila, C1+j, r[col])
        for c in range(C1, C1+5):
            ws.cell(fila, c).fill = FILLS.get(r['CONFIANZA'], PatternFill())
    for c in range(C1, C1+5):
        ws.column_dimensions[get_column_letter(c)].width = 30
    # RESUMEN
    total = len(resultados)
    if 'RESUMEN' in wb.sheetnames: del wb['RESUMEN']
    rs = wb.create_sheet('RESUMEN')
    rs.append(['CONCILIACION BANCARIA — EMETE Asesoria',''])
    rs.append(['Generado', datetime.now().strftime('%d/%m/%Y %H:%M')])
    rs.append(['Total', total]); rs.append(['',''])
    rs.append(['CONFIANZA','N','%'])
    for conf in ['ALTA','MEDIA','BAJA','REVISAR']:
        n = sum(1 for r in resultados if r['CONFIANZA']==conf)
        rs.append([conf, n, f"{n/total*100:.1f}%"])
    rs.append(['','','']); rs.append(['TOP CUENTAS','N',''])
    from collections import Counter
    for (cod,desc),n in Counter((r['CUENTA'],r['DESCRIPCION'][:35]) for r in resultados)\
                         .most_common(20):
        rs.append([f"{cod}  {desc}", n, ''])
    rs.column_dimensions['A'].width = 55
    # REVISAR
    if 'REVISAR' in wb.sheetnames: del wb['REVISAR']
    rv = wb.create_sheet('REVISAR')
    rv.append(['F_CONTABLE','CONCEPTO','BENEFICIARIO','OBSERVACIONES','IMPORTE',
               'CUENTA','DESCRIPCION','CONFIANZA','JUSTIFICACION'])
    for r in resultados:
        if r['CONFIANZA'] in ('REVISAR','BAJA'):
            row = df.iloc[r['idx']]
            rv.append([row['F_CONTABLE'],row['CONCEPTO'],row['BENEFICIARIO'],
                       row['OBSERVACIONES'],row['IMPORTE'],
                       r['CUENTA'],r['DESCRIPCION'],r['CONFIANZA'],r['JUSTIFICACION']])
    for col in 'ABCDEFGHI': rv.column_dimensions[col].width = 26
    _hoja_cegid(wb, df, resultados)
    if acciones:
        _hoja_acciones(wb, acciones)
    wb.save(output_path)
    print(f"Guardado: {output_path}")

# ── SELF-CHECK (verificación de instalación, sin ficheros) ────────────────────
def _self_check():
    # Parser de importes: formatos español, US y negativos
    assert _smart_float("1.234,56") == 1234.56
    assert _smart_float("-45,00") == -45.0
    assert _smart_float("1,234.56") == 1234.56
    assert _smart_float("€ 900,00") == 900.0
    assert _smart_float(None) == 0.0
    # Cascada de clasificación (sin terceros en LM)
    init_cache_lm({})
    def fila(conc='', obs='', benef='', imp=-100.0):
        return {'CONCEPTO':conc,'OBSERVACIONES':obs,'BENEFICIARIO':benef,'IMPORTE':imp}
    assert clasificar_fila(fila(obs='NOMINA ENERO'), {})[0] == '46500000'
    assert clasificar_fila(fila(obs='CUOTA SS'), {})[0] == '47600000'
    assert clasificar_fila(fila(conc='COMISION MANTENIMIENTO CUENTA'), {})[0] == '62600000'
    assert clasificar_fila(fila(obs='RETIRADA EFECTIVO CAJERO'), {})[0] == '57000000'
    assert clasificar_fila(fila(benef='PROVEEDOR DESCONOCIDO SL'), {})[0] == '41000000'
    # Regla de oro: abono de comisión → ingreso, no gasto
    assert clasificar_fila(fila(conc='COMISION MANTENIMIENTO CUENTA', imp=5.0), {})[0] == '76900000'
    # Task 1: cargar_pgc indexa por 3 y por 8 dígitos
    from openpyxl import Workbook as _WB
    import tempfile, os as _os
    _wb = _WB(); _ws = _wb.create_sheet("Datos"); del _wb[_wb.sheetnames[0]]
    # fila 6 en adelante: col índice 4 = código 3 díg, col 8 = subcuenta 8 díg, col 9 = nombre
    for _ in range(5): _ws.append([])
    _ws.append([None,None,None,None,"572",None,None,None,"57200000","BANCOS"])
    _tmp = _os.path.join(tempfile.gettempdir(), "_pgc_test.xlsx"); _wb.save(_tmp)
    _pgc = cargar_pgc(_tmp); _os.remove(_tmp)
    assert _pgc.get("572") == "BANCOS"
    assert _pgc.get("57200000") == "BANCOS"
    # Task 2: toks separa camelCase/pegados
    assert "CLIENTES" in toks("TotalEnergiesClientesSAU")
    assert "ENERGIES" in toks("TotalEnergiesClientesSAU")
    # Task 3: alias de criterios gana al matcher por nombre
    _crit = {'alias_terceros': {'FACEBK': '41000008'}}
    _r = clasificar_fila(fila(obs='PAGO FACEBK ADS'), {}, _crit)
    assert _r[0] == '41000008' and _r[4] == 'ALIAS', f"Expected ('41000008', ..., 'ALIAS'), got {_r}"
    # cargar_criterios devuelve {} si no hay fichero
    _tmp_dir = tempfile.gettempdir() + '/no_existe_dir_xyz'
    assert cargar_criterios(_tmp_dir) == {}, f"Expected {{}}, got {cargar_criterios(_tmp_dir)}"
    # Task 4: match por importe único → ALTA; múltiple → no asigna
    init_importes_lm({'40000001': {125.84}, '40000002': {99.0}})
    _r = clasificar_fila(fila(benef='PROV RARO', imp=-125.84), {'40000001': 'FLOCCUS'})
    assert _r[0] == '40000001' and _r[4] == 'MATCH_IMPORTE'
    init_importes_lm({'40000001': {125.84}, '40000002': {125.84}})
    _r = clasificar_fila(fila(benef='PROV RARO', imp=-125.84), {'40000001': 'A', '40000002': 'B'})
    assert _r[4] != 'MATCH_IMPORTE'   # ambiguo → no adivina
    init_importes_lm({})              # limpiar para no afectar otros asserts
    # Task 5: TPV usa cuenta de criterios; nómina con texto intermedio; SS truncada
    _r = clasificar_fila(fila(conc='LIQUIDACION REMESA DE COMERCIOS', imp=1500.0),
                         {}, {'cuenta_ventas_tpv': '70500000'})
    assert _r[0] == '70500000'
    assert clasificar_fila(fila(obs='NOMINA LIQUIDACION DE JUNIO 2025'), {})[0] == '46500000'
    assert clasificar_fila(fila(obs='TESORERIA GENERAL DE LA SEGURID'), {})[0] == '47600000'
    # Task 6: par +/- retrocedido → 555
    _df = pd.DataFrame([
        {'F_CONTABLE':'01/03/2025','F_VALOR':'','CODIGO':'','CONCEPTO':'CONFIRMING',
         'BENEFICIARIO':'','OBSERVACIONES':'','IMPORTE':300.0,'SALDO':0.0},
        {'F_CONTABLE':'02/03/2025','F_VALOR':'','CODIGO':'','CONCEPTO':'RETROCESION CONFIRMING',
         'BENEFICIARIO':'','OBSERVACIONES':'','IMPORTE':-300.0,'SALDO':0.0},
    ])
    _res = [{'idx':0,'CUENTA':'X','DESCRIPCION':'','CONFIANZA':'BAJA','JUSTIFICACION':'','METODO':'x'},
            {'idx':1,'CUENTA':'Y','DESCRIPCION':'','CONFIANZA':'BAJA','JUSTIFICACION':'','METODO':'y'}]
    marcar_retrocesiones(_df, _res)
    assert _res[0]['CUENTA'] == '55500000' and _res[1]['CUENTA'] == '55500000'
    # Task 7: validar_saldo detecta roturas
    _ok = pd.DataFrame([{'IMPORTE':-10.0,'SALDO':90.0},{'IMPORTE':-5.0,'SALDO':100.0}])
    assert validar_saldo(_ok) == []          # 90 == 100 + (-10)
    _bad = pd.DataFrame([{'IMPORTE':-10.0,'SALDO':80.0},{'IMPORTE':-5.0,'SALDO':100.0}])
    assert validar_saldo(_bad) == [0]        # 80 != 100 + (-10)
    # Task 8: parser PDF sobre líneas sintéticas BBVA (concepto|beneficiario arriba, obs abajo)
    _lineas = [
        (0, 10.0, "PAGO NOMINA | JUAN PEREZ"),
        (0, 20.0, "05/06/2025 05/06/2025 1234 -1.200,00 EUR 3.800,00 EUR"),
        (0, 30.0, "TRANSFERENCIA NOMINA JUNIO"),
        (0, 60.0, "COMPRA TARJETA | MERCADONA"),           # gap 30 > 18 → nuevo movimiento
        (0, 70.0, "06/06/2025 06/06/2025 5678 -45,20 EUR 3.754,80 EUR"),
    ]
    _movs = _parsear_movimientos_pdf(_lineas)
    assert len(_movs) == 2
    assert _movs[0]['IMPORTE'] == -1200.0 and _movs[0]['BENEFICIARIO'] == 'JUAN PEREZ'
    assert _movs[0]['CONCEPTO'] == 'PAGO NOMINA'
    assert _movs[1]['IMPORTE'] == -45.2 and _movs[1]['CONCEPTO'] == 'COMPRA TARJETA'
    # Task 9: hoja CEGID con 4 columnas exactas
    from openpyxl import Workbook as _WB2
    _wb2 = _WB2()
    _df2 = pd.DataFrame([{'F_CONTABLE':'05/06/2025','CONCEPTO':'PAGO','IMPORTE':-10.0,
                          'BENEFICIARIO':'','OBSERVACIONES':'','SALDO':0.0}])
    _res2 = [{'idx':0,'CUENTA':'41000001','DESCRIPCION':'','CONFIANZA':'ALTA',
              'JUSTIFICACION':'','METODO':'x'}]
    _hoja_cegid(_wb2, _df2, _res2)
    _ws2 = _wb2['CEGID']
    assert [c.value for c in _ws2[1]] == ['Fecha','Concepto','Importe','Contrapartida']
    assert [c.value for c in _ws2[2]] == ['05/06/2025','PAGO',-10.0,'41000001']
    # Task 10: siguiente código libre por bloque
    assert siguiente_codigo_libre({'40000001':'A','40000005':'B'}, '400') == '40000006'
    assert siguiente_codigo_libre({}, '410') == '41000001'
    # detectar_acciones agrupa genéricos y propone código; detecta subcuenta inexistente
    _df3 = pd.DataFrame([
        {'F_CONTABLE':'01/06/2025','CONCEPTO':'PAGO','BENEFICIARIO':'FERRETERIA LOPEZ',
         'OBSERVACIONES':'FERRETERIA LOPEZ','IMPORTE':-80.0,'SALDO':0.0},
    ])
    _res3 = [{'idx':0,'CUENTA':'41000000','DESCRIPCION':'','CONFIANZA':'BAJA',
              'JUSTIFICACION':'','METODO':'ACREEDOR_GENERICO'}]
    _acc = detectar_acciones(_df3, _res3, {}, {})
    assert len(_acc['terceros_crear']) == 1
    assert _acc['terceros_crear'][0]['codigo'] == '41000001'
    # Task 11: hoja ACCIONES_CEGID con los 3 bloques
    from openpyxl import Workbook as _WB3
    _wb3 = _WB3()
    _acc3 = {'terceros_crear':[{'codigo':'41000001','nombre':'FERRETERIA LOPEZ',
              'tipo':'acreedor','nif':'(pendiente)','n_movimientos':2}],
             'subcuentas_crear':[{'codigo':'62800000','nombre':'SUMINISTROS'}],
             'avisos':[{'idx':3,'motivo':'Sin beneficiario','importe':-9.0}]}
    _hoja_acciones(_wb3, _acc3)
    _wsa = _wb3['ACCIONES_CEGID']
    _textos = [str(c.value) for r in _wsa.iter_rows() for c in r if c.value]
    assert any('FERRETERIA LOPEZ' in t for t in _textos)
    assert any('62800000' in t for t in _textos)
    assert any('TERCEROS A CREAR' in t.upper() for t in _textos)
    print("self-check OK")

# ── MAIN ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    from datetime import date as _date
    # Sin los tres ficheros presentes → self-check (verificación de instalación).
    if not all(os.path.exists(p) for p in (EXTRACTO_PATH, LM_PATH, PGC_PATH)):
        _self_check()
    else:
        fd = _date.fromisoformat(FECHA_DESDE) if FECHA_DESDE else None
        fh = _date.fromisoformat(FECHA_HASTA) if FECHA_HASTA else None
        print("Cargando extracto...")
        df = cargar_extracto(EXTRACTO_PATH, fd, fh)
        print(f"  {len(df)} movimientos")
        print("Cargando libro mayor...")
        terceros = cargar_lm(LM_PATH)
        print(f"  {len(terceros)} terceros (400/410/430)")
        print("Cargando plan de cuentas...")
        pgc = cargar_pgc(PGC_PATH)
        print(f"  {len(pgc)} cuentas")
        print("Clasificando...")
        res = conciliar(df, terceros, pgc, None, API_KEY)
        print("Generando Excel...")
        generar_excel(df, res, EXTRACTO_PATH, OUTPUT_PATH)
        total = len(res)
        print("\nRESULTADO:")
        for conf in ['ALTA','MEDIA','BAJA','REVISAR']:
            n = sum(1 for r in res if r['CONFIANZA']==conf)
            print(f"  {conf:<10} {n:>4}  ({n/total*100:.1f}%)")
